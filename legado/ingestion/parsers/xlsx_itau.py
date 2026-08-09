import warnings
from datetime import date, datetime
from decimal import Decimal

import openpyxl

from ..schemas import ResultadoParse, Snapshot, Transacao

warnings.filterwarnings("ignore", module="openpyxl")

_COL = {"data": 1, "lancamento": 2, "parcelamento": 3, "valor": 4}
_PAGAMENTO = ("PAGAMENTO", "PAGTO")


def _achar_header(rows) -> int:
    """A planilha começa com metadados do titular; o header real está no meio."""
    for i, r in enumerate(rows):
        celulas = [str(c).strip() if c else "" for c in r]
        if "Data" in celulas and "Lançamento" in celulas:
            return i
    raise ValueError("Header de lançamentos não encontrado na planilha")


def _achar_valor_fatura(rows):
    for i, r in enumerate(rows):
        if any(c and "Valor (parcial)" in str(c) for c in r):
            for c in rows[i + 1]:
                if isinstance(c, (int, float)):
                    return Decimal(str(c))
    return None


def parse(caminho: str, conta_id: int, fonte: str) -> ResultadoParse:
    """Lê a fatura do Itaú em XLSX.

    ATENÇÃO AO SINAL: na planilha do Itaú compras são POSITIVAS (quanto você
    deve) e pagamentos são NEGATIVOS. Isso é o inverso da convenção do projeto,
    então todo valor é multiplicado por -1.
    """
    res = ResultadoParse()

    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    rows = [r for r in wb[wb.sheetnames[0]].iter_rows(values_only=True)]

    inicio = _achar_header(rows) + 1

    for r in rows[inicio:]:
        bruto_data = r[_COL["data"]] if len(r) > _COL["data"] else None
        # Linhas de "Subtotal" e o rodapé de avisos param a leitura.
        if not isinstance(bruto_data, (datetime, date)):
            if any(c and "Subtotal" in str(c) for c in r if c):
                break
            continue

        lancamento = r[_COL["lancamento"]]
        bruto_valor = r[_COL["valor"]]
        if lancamento is None or bruto_valor is None:
            continue

        valor = Decimal(str(bruto_valor)) * -1  # inversão de sinal
        d = bruto_data.date() if isinstance(bruto_data, datetime) else bruto_data
        eh_pagamento = any(p in str(lancamento).upper() for p in _PAGAMENTO)

        res.transacoes.append(
            Transacao(
                conta_id=conta_id,
                data=d,
                valor=valor,
                descricao=str(lancamento),
                fonte=fonte,
                metodo="credito",
                eh_interna=eh_pagamento,
            )
        )

    # "Valor (parcial)" é a soma das COMPRAS do ciclo, e não o saldo líquido:
    # ele ignora pagamentos lançados no período. Emitimos como snapshot (negativo,
    # passivo) mas avisamos, porque fechar a conferência exige um lançamento de
    # saldo de abertura do ciclo.
    valor_fatura = _achar_valor_fatura(rows)
    if valor_fatura is not None:
        soma = sum(t.valor for t in res.transacoes)
        res.snapshot = Snapshot(
            conta_id=conta_id,
            data_ref=max(t.data for t in res.transacoes),
            saldo=valor_fatura * -1,
            fonte=fonte,
            observacao="Valor (parcial) da fatura: soma de compras, não saldo líquido",
        )
        abertura = (valor_fatura * -1) - soma
        if abertura != 0:
            res.avisos.append(
                f"Saldo de abertura do ciclo necessário: {abertura:.2f} "
                f"(compras {soma:.2f} vs fatura {valor_fatura * -1:.2f})"
            )

    return res
